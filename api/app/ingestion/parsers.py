"""CSV and XLSX to raw rows. Nothing here interprets a value -- it only reads cells.

Deliberately no pandas: it type-coerces on read, which would silently reinterpret exactly the
strings the normalizer must inspect verbatim (``03/04/2026`` must survive as text so the batch
date-format detector can rule on it, not arrive pre-guessed as a datetime).

The one exception is a *genuine* Excel date cell. Excel stores those as numbers with a date
format, carrying no DD/MM ambiguity at all, so openpyxl handing back a ``date`` is strictly more
information than a string. Those are emitted as ISO ``YYYY-MM-DD``, which the date parser reads
unambiguously. See :func:`_cell_to_text`.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from app.exceptions import IngestionError

# Guardrail against a malicious or accidental multi-hundred-MB upload.
MAX_ROWS = 50_000


@dataclass(frozen=True)
class ParsedFile:
    """Raw tabular content. ``rows`` are header -> cell text, in file order."""

    headers: list[str]
    rows: list[dict[str, str]] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)


def _cell_to_text(value: Any) -> str:
    """Render one cell as text without guessing at its meaning.

    A real ``date``/``datetime`` (an Excel date-formatted cell) becomes ISO, because that is the
    unambiguous rendering of information Excel already resolved for us. Everything else is
    stringified as-is; ``None`` becomes an empty string.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # openpyxl returns numerics as float; 410000.0 must not reach the amount parser as
        # "410000.0" when the file said "410000".
        return str(int(value))
    return str(value).strip()


def _dedupe_headers(headers: list[str]) -> list[str]:
    """Make headers unique and non-empty so a dict row cannot silently lose a column."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for index, header in enumerate(headers):
        name = header.strip() or f"column_{index + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 0
        out.append(name)
    return out


def _rows_from_matrix(matrix: list[list[Any]]) -> ParsedFile:
    non_empty = [row for row in matrix if any(_cell_to_text(c) for c in row)]
    if not non_empty:
        raise IngestionError("file contains no rows")
    headers = _dedupe_headers([_cell_to_text(c) for c in non_empty[0]])
    body = non_empty[1:]
    if len(body) > MAX_ROWS:
        raise IngestionError(f"file has {len(body)} rows; the limit is {MAX_ROWS}")

    rows: list[dict[str, str]] = []
    for raw_row in body:
        cells = [_cell_to_text(c) for c in raw_row]
        # Pad/truncate to the header width so every row has the same keys.
        cells += [""] * (len(headers) - len(cells))
        rows.append(dict(zip(headers, cells[: len(headers)], strict=True)))
    return ParsedFile(headers=headers, rows=rows)


def parse_csv(data: bytes) -> ParsedFile:
    """Parse CSV/TSV bytes. Sniffs the delimiter; falls back to comma."""
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        # Tally and older Excel exports on Indian Windows are frequently cp1252.
        text = data.decode("cp1252", errors="replace")

    sample = text[:8192]
    try:
        dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    matrix = [list(row) for row in csv.reader(io.StringIO(text), dialect)]
    return _rows_from_matrix(matrix)


def parse_xlsx(data: bytes) -> ParsedFile:
    """Parse the first worksheet of an XLSX workbook."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a wide range on malformed input
        raise IngestionError(f"could not read workbook: {exc}") from exc
    try:
        sheet = workbook.worksheets[0]
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return _rows_from_matrix(matrix)


def parse_upload(filename: str, data: bytes) -> ParsedFile:
    """Dispatch on file extension. Raises :class:`IngestionError` for anything else."""
    lowered = (filename or "").lower()
    if lowered.endswith((".csv", ".tsv", ".txt")):
        return parse_csv(data)
    if lowered.endswith((".xlsx", ".xlsm")):
        return parse_xlsx(data)
    if lowered.endswith(".xls"):
        raise IngestionError("legacy .xls is not supported; re-save as .xlsx or .csv")
    raise IngestionError(f"unsupported file type: {filename!r}; expected .csv or .xlsx")
